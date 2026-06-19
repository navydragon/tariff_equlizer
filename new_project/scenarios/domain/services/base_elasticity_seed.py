"""Начальные значения эластичности угля из IPEM (лист Уголь_коэфф)."""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from django.contrib.auth import get_user_model
from django.db import transaction

from core.models import MessageType
from scenarios.models import ElasticityRule, ElasticityRulePoint, ElasticitySet, Scenario

User = get_user_model()

ELASTICITY_SET_NAME = "2026"
EXPORT_RULE_NAME = "Уголь экспорт"
INTERNAL_RULE_NAME = "Уголь внутренние"

SEEDED_RULE_NAMES = (EXPORT_RULE_NAME, INTERNAL_RULE_NAME)


@dataclass(frozen=True)
class ElasticitySeedResult:
    elasticity_set_id: int
    points_export: int
    points_internal: int
    attached_to_scenario: bool

    @property
    def points_upserted(self) -> int:
        return self.points_export + self.points_internal


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[4]


def _load_coefficient_points(
    worksheet,
    *,
    marginality_col: int,
    coefficient_col: int,
) -> list[tuple[Decimal, Decimal]]:
    by_marginality: dict[Decimal, Decimal] = {}
    for row in range(3, worksheet.max_row + 1):
        marginality = worksheet.cell(row, marginality_col).value
        coefficient = worksheet.cell(row, coefficient_col).value
        if marginality is None or coefficient is None:
            continue
        if str(marginality).strip() == "" or str(coefficient).strip() == "":
            continue
        key = Decimal(str(marginality)).quantize(Decimal("0.0001"))
        by_marginality[key] = Decimal(str(coefficient)).quantize(Decimal("0.0001"))
    return sorted(by_marginality.items(), key=lambda item: item[0])


def _load_coal_workbook_points(
    xlsx_path: Path | None = None,
) -> tuple[list[tuple[Decimal, Decimal]], list[tuple[Decimal, Decimal]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to seed coal elasticity") from exc

    workbook_path = xlsx_path or (_repo_root() / "data" / "ipem" / "Уголь_эластика_2026.xlsx")
    if not workbook_path.exists():
        return [], []

    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    worksheet = workbook["Уголь_коэфф"]
    export_points = _load_coefficient_points(
        worksheet,
        marginality_col=1,
        coefficient_col=2,
    )
    internal_points = _load_coefficient_points(
        worksheet,
        marginality_col=5,
        coefficient_col=6,
    )
    return export_points, internal_points


def _find_message_type_by_keyword(keyword: str) -> MessageType | None:
    return (
        MessageType.objects.filter(name__icontains=keyword)
        .order_by("id")
        .first()
    )


def _replace_rule_points(
    rule: ElasticityRule,
    points: list[tuple[Decimal, Decimal]],
) -> int:
    ElasticityRulePoint.objects.filter(rule=rule).delete()
    if not points:
        return 0
    ElasticityRulePoint.objects.bulk_create(
        [
            ElasticityRulePoint(
                rule=rule,
                marginality=marginality,
                coefficient=coefficient,
            )
            for marginality, coefficient in points
        ],
    )
    return len(points)


def _resolve_elasticity_set_for_seed(owner: User) -> ElasticitySet:
    elasticity_set = (
        ElasticitySet.objects.filter(name=ELASTICITY_SET_NAME)
        .order_by("id")
        .first()
    )
    if elasticity_set is None:
        return ElasticitySet.objects.create(
            author=owner,
            name=ELASTICITY_SET_NAME,
        )
    ElasticityRule.objects.filter(elasticity_set=elasticity_set).delete()
    return elasticity_set


@transaction.atomic
def seed_coal_elasticity_for_scenario(
    scenario: Scenario,
    *,
    author: User | None = None,
    attach: bool = True,
    xlsx_path: Path | None = None,
) -> ElasticitySeedResult:
    owner = author or scenario.author
    if owner is None:
        raise ValueError("author is required to seed elasticity set")

    elasticity_set = _resolve_elasticity_set_for_seed(owner)

    export_points, internal_points = _load_coal_workbook_points(xlsx_path)
    export_message_type = _find_message_type_by_keyword("экспорт")
    internal_message_type = _find_message_type_by_keyword("внутр")

    export_rule = ElasticityRule.objects.create(
        elasticity_set=elasticity_set,
        name=EXPORT_RULE_NAME,
        position=0,
        message_type=export_message_type,
    )
    internal_rule = ElasticityRule.objects.create(
        elasticity_set=elasticity_set,
        name=INTERNAL_RULE_NAME,
        position=1,
        message_type=internal_message_type,
    )
    export_count = _replace_rule_points(export_rule, export_points)
    internal_count = _replace_rule_points(internal_rule, internal_points)

    attached = False
    if attach and not scenario.elasticity_set_id:
        scenario.elasticity_set = elasticity_set
        scenario.save(update_fields=["elasticity_set"])
        attached = True

    return ElasticitySeedResult(
        elasticity_set_id=elasticity_set.id,
        points_export=export_count,
        points_internal=internal_count,
        attached_to_scenario=attached,
    )
